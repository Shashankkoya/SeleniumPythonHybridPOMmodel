import time

from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("--incognito")
driver = webdriver.Chrome(options=chrome_options)
# Load the Excel sheet

workbook=load_workbook('testdata.xlsx')
# Selecting the sheet
sheet=workbook.active

driver.maximize_window()
time.sleep(3)

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/")
    time.sleep(5)
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    time.sleep(3)

    driver.find_element(By.XPATH,"//button[@id='react-burger-menu-btn']").click()
    time.sleep(3)

    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(3)
